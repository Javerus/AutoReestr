from tkinter import filedialog
import tkinter as tk
import pandas as pd
import openpyxl
import itertools

# Переменные для записи данных
Namber_Acts, Data_Acts, Named_work, IS, materials = [], [], [], [], []


""" модуль: работа с файлами"""


def open_read_files():
    # Открываем файл чтения
    file_path_o = filedialog.askopenfilename()
    print('Открываем файл чтения', file_path_o)
    return file_path_o


def open_file_write():
    # Открываем файл записи
    file_path_w = filedialog.askopenfilename()
    print('Открываем файл записи', file_path_w)
    return file_path_w


def read_files(file_path_o):
    # Читаем данные
    path_file_read_and_write_excel = file_path_o
    return path_file_read_and_write_excel


def open_book_write(file_path_w):
    # Открываем книгу для записи данных
    path_file_write_excel = file_path_w
    workbook = openpyxl.load_workbook(path_file_write_excel)
    return workbook


def reading_data(path_file_read_and_write_excel):
    # читаем лист
    df = pd.read_excel(path_file_read_and_write_excel, sheet_name='БД (5)')
    return df


def open_list_write(workbook):
    # выбираем лист, в который будем записывать значения
    worksheet = workbook['Реестр']
    return worksheet


def show_sheet(sheet_name, root):
    # Функция, которая будет вызываться при нажатии на кнопку листа
    print(f"Имя выбранного листа: {sheet_name}")
    root.destroy()


def get_sheet_name():
    # Открытие книги Excel
    wb = openpyxl.load_workbook('/Users/gregoryreyn/Desktop/Общая БД 2.xlsm')
    # Создание окна с кнопками листов
    root = tk.Tk()
    root.title("Выберите лист")
    for i, sheet in enumerate(wb.sheetnames):
        # Создание кнопки с именем листа
        btn = tk.Button(root, text=sheet, command=lambda sheet_name=sheet: show_sheet(sheet_name, root))
        btn.grid(row=0, column=i)
    root.mainloop()
    return root


""" модуль: логика"""


def parsing_data(df):
    # парсим и записываем данные в переменные
    for index, row in df.iterrows():
        Namber_Acts.append(row['№ акта'])
        Data_Acts.append(row['дата акта'])
        Named_work.append(row['1. К освидетельствованию предъявлены следующие работы:'])
        IS.append(row['4. Предъявлены документы'])
        materials.append(row['3. При выполнении работ применены:'])


def act_gen():
    # генерируем акты
    aosr, act_namber, act_name_work = 'АОСР', Namber_Acts, Named_work
    finel_act = []
    for i in range(len(act_name_work)):
        finel_act.append(f"{aosr} {act_namber[i]} {act_name_work[i]}")
    return finel_act


def is_gen():
    # генерируем схемы (разделяем схемы)
    my_list1 = IS
    finel_IS = []
    for item in my_list1:
        split_str = item.split(";")
        finel_IS += [x.strip() for x in split_str]
    return finel_IS


def materials_gen():
    # генерируем паспорта/сертификаты (разделяем паспорта/сертификаты)
    materials_list = materials
    finel_materials = [materials]
    for item in materials_list:
        split_str = item.split(";")
        finel_materials += [x.strip() for x in split_str]
    pass


def osi_gen(finel_act):
    finel_osi_otm = []  # новый список, куда будут записываться значения
    for item in finel_act:
        if "в осях" in item:
            new_item = item.split("в осях", 1)[1].strip()
            finel_osi_otm.append("в осях " + new_item)
    return finel_osi_otm


def all_write(file_path_w, worksheet, workbook, finel_act, finel_IS, finel_osi_otm):
    starting_row = 12  # определяем ячейку, с которой начнём записывать значения
    starting_column = 3
    for write_finel_act, write_finel_IS, write_finel_materials, write_data_Acts, write_finel_osi_otm in itertools.zip_longest(
            finel_act, finel_IS, materials, Data_Acts, finel_osi_otm, fillvalue="-"):
        worksheet.cell(row=starting_row, column=starting_column, value=write_finel_act)
        worksheet.cell(row=starting_row, column=starting_column + 1, value=write_finel_osi_otm)
        worksheet.cell(row=starting_row, column=starting_column + 2, value=write_data_Acts)
        worksheet.cell(row=starting_row + 1, column=starting_column, value=write_finel_IS)
        worksheet.cell(row=starting_row + 2, column=starting_column, value=write_finel_materials)
        starting_row += 3
    workbook_save = workbook.save(file_path_w)
    # Проверка финальных данных
    print('финальные акты:', finel_act)
    print('Дата акта:', Data_Acts)
    print('финальные оси и отметки:', finel_osi_otm)
    print('финальые Исполнительные схемы:', finel_IS)
    print('материалы:', materials)
    return workbook_save


""" модуль: запуск"""


def main():
    # вызываем функцию open_read_files()
    file_path_o = open_read_files()
    # вызываем функцию open_file_write()
    file_path_w = open_file_write()
    # вызываем функцию read_files()
    path_file_read_and_write_excel = read_files(file_path_o)
    # вызываем функцию open_book_write()
    workbook = open_book_write(file_path_w)
    # вызываем функцию reading_data()
    df = reading_data(path_file_read_and_write_excel)
    # вызываем функцию open_list_write()
    worksheet = open_list_write(workbook)
    # вызываем функцию parsing_data()
    parsing_data(df)
    # вызываем функцию act_gen()
    finel_act = act_gen()
    # вызываем функцию is_gen()
    finel_IS = is_gen()
    # вызываем функцию materials_gen()
    materials_gen()
    # вызываем функцию osi_gen()
    finel_osi_otm = osi_gen(finel_act)
    # вызываем функцию all_write()
    all_write(file_path_w, worksheet, workbook, finel_act, finel_IS, finel_osi_otm)


""" модуль: интерфейс"""


def button():
    root = tk.Tk()
    root.title("кря")
    button1 = tk.Button(root, text="Открыть файл чтения", command='')
    button1.pack()
    button2 = tk.Button(root, text="Открыть файл записи", command='')
    button2.pack()
    button3 = tk.Button(root, text="запись", command=lambda: main())
    button3.pack()
    button3 = tk.Button(root, text="проверка кнопок", command=get_sheet_name)
    button3.pack()
    root.mainloop()


button()
