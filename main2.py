import pandas as pd
import openpyxl
import itertools


# Переменные для записи данных
namber_acts, data_acts, named_work, IS, materials = [], [], [], [], []
finel_act, finel_IS, finel_materials, finel_osi_otm, finel_data_acts = [], [], [], [], []
data_acts = finel_data_acts


def path_file():
    # Определяем путь
    path_file_read_and_write_excel = '/Users/gregoryreyn/Desktop/Общая БД.xlsm'
    return path_file_read_and_write_excel


def reading_data(path_file_read_and_write_excel):
    # Открываем книгу для чтения
    df = pd.read_excel(path_file_read_and_write_excel, sheet_name='БД')
    return df


def open_book_write():
    # открываем книгу для записи данных
    path_file_write_excel = '/Users/gregoryreyn/Desktop/тест книга.xlsx'  # заменить на выбор
    workbook = openpyxl.load_workbook(path_file_write_excel)
    return workbook


def open_list_write(workbook):
    # выбираем лист, в который будем записывать значения
    worksheet = workbook['Лист2']  # заменить на выбор
    return worksheet


def parsing_data(df):
    # парсим и записываем данные в переменные
    for index, row in df.iterrows():
        namber_acts.append(row['№ акта'])
        data_acts.append(row['дата акта'])
        named_work.append(row['1. К освидетельствованию предъявлены следующие работы:'])
        IS.append(row['4. Предъявлены документы'])
        materials.append(row['3. При выполнении работ применены:'])
    return df


def act_gen():
    aosr, act_namber, act_name_work = 'АОСР', namber_acts, data_acts
    for i in range(len(act_name_work)):
        finel_act.append(f"{aosr} {act_namber[i]} {act_name_work[i]}")


def is_gen():
    my_list1 = IS
    finel_IS_l = finel_IS
    for item in my_list1:
        split_str = item.split(";")
        finel_IS_l += [x.strip() for x in split_str]
    finel_IS_l = finel_IS
    return finel_IS


# def materials_gen():
#     # генерируем паспорта/сертификаты (разделяем паспорта/сертификаты)
#     materials_list = materials
#     finel_materials = [materials]
#     for item in materials_list:
#         split_str = item.split(";")
#         finel_materials += [x.strip() for x in split_str]
#     pass


def osi_gen():
    # finel_osi_otm = []  # новый список, куда будут записываться значения
    for item in finel_act:
        if "в осях" in item:
            new_item = item.split("в осях", 1)[1].strip()
            finel_osi_otm.append("в осях " + new_item)


def all_write(worksheet):
    # записываем каждое значение из списка в ячейку
    starting_row = 12  # определяем ячейку, с которой начнём записывать значения
    starting_column = 3
    # заполняем пропущенные значения значением "-"
    for write_finel_act, write_finel_IS, write_finel_materials, write_finel_data_acts, write_finel_osi_otm in itertools.zip_longest(
            finel_act, finel_IS, materials, finel_data_acts, finel_osi_otm, fillvalue="-"):
        worksheet.cell(row=starting_row, column=starting_column, value=write_finel_act)
        worksheet.cell(row=starting_row, column=starting_column + 1, value=write_finel_osi_otm)
        worksheet.cell(row=starting_row, column=starting_column + 2, value=write_finel_data_acts)
        worksheet.cell(row=starting_row + 1, column=starting_column, value=write_finel_IS)
        worksheet.cell(row=starting_row + 2, column=starting_column, value=write_finel_materials)
        starting_row += 3


def main():
    """This function is the main function that calls other functions"""
    p1 = path_file()
    p2 = reading_data(p1)
    p3 = parsing_data(p2)
    print(namber_acts, data_acts, named_work, IS, materials)
    return p1, p2, p3


main()
